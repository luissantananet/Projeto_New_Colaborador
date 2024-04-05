import os
import sqlite3
import openpyxl
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from PyQt5 import uic, QtWidgets
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem
from PIL import Image, ImageDraw, ImageFont
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4


# Criando o Bando de Dados
# Caminho para a pasta onde o banco de dados será armazenado
pasta_dados = r'.\dados'
# Verifica se a pasta "dados" existe, se não, cria a pasta
if not os.path.exists(pasta_dados):
    os.makedirs(pasta_dados)
# Caminho completo para o arquivo do banco de dados
database = os.path.join(pasta_dados, 'banco_cadastro.db')
# Verifica se o arquivo do banco de dados já existe
if not os.path.exists(database):
    # Conecta ao banco de dados (isso criará o arquivo se ele não existir)
    banco = sqlite3.connect(database)
    # Fechando a conexão
    banco.close()
# criando tabelas se ele nao exixtir
banco = sqlite3.connect(database)
cursor = banco.cursor()
# Cria a tabela 'tabela' se ela não existir
cursor.execute("""CREATE TABLE IF NOT EXISTS tabela ( 
id INTEGER PRIMARY KEY AUTOINCREMENT, 
diaria decimal(5,2) NOT NULL, 
hextra decimal(5,2) NOT NULL, 
vtransp decimal(5,2) NOT NULL,
vref decimal(5,2) NOT NULL);""")
# Cria a tabela 'funcoes' se ela não existir
cursor.execute("""CREATE TABLE IF NOT EXISTS funcoes (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    descricao TEXT NOT NULL);""")
# Cria a tabela 'cadastro_colaborador' se ela não existir
cursor.execute("""CREATE TABLE IF NOT EXISTS cadastro_colaborador ( 
        id INTEGER PRIMARY KEY AUTOINCREMENT, 
        nome_completo TEXT NOT NULL, 
        funcao TEXT NOT NULL, 
        cpf TEXT NOT NULL,
        rg TEXT, 
        cnh TEXT);""")
# Salva as alterações no banco de dados
banco.commit()
# Fecha a conexão com o banco de dados
banco.close()


def gerarPlanilha():
    nome_arquivo = 'RegistrosColaboradores.xlsx'
    arquivo_xlsx = os.path.join(pasta_dados, nome_arquivo)

    # Verifica se o diretório 'dados' existe, se não, cria o diretório
    if not os.path.exists(pasta_dados):
        os.makedirs(pasta_dados)

    # Verifica se o arquivo já existe no diretório 'dados'
    if not os.path.isfile(arquivo_xlsx):
        # Criando um novo livro
        wb = Workbook()
        # Renomeando a aba padrão para 'Registros'
        wsRegistro = wb.active
        wsRegistro.title = "Registros"
        # Criando e renomeando as outras abas
        wsColaboradores = wb.create_sheet("Colaboradores")
        wsFuncoes = wb.create_sheet("Funções")
        wsTaxas = wb.create_sheet("Taxas")
        # Salvar no diretório 'dados'
        wb.save(arquivo_xlsx)
        QMessageBox.information(
            frm_principal, "Aviso", f"O arquivo {nome_arquivo} foi criado com sucesso no diretório {pasta_dados}.")
    else:
        QMessageBox.information(
            frm_principal, "Aviso", f"O arquivo {nome_arquivo} já existe no diretório {pasta_dados}.")


def salvarRegistro():
    # Caminho para o arquivo existente
    arquivos_xlsx = r'.\dados\RegistrosColaboradores.xlsx'

    # Carrega o workbook existente
    wb = load_workbook(arquivos_xlsx)

    # Seleciona a aba 'Registros'
    if 'Registros' in wb.sheetnames:
        ws = wb['Registros']
    else:
        # Se a aba 'Registros' não existir, cria uma nova aba
        ws = wb.create_sheet('Registros')

    # Obtém os dados do formulário
    data_inicial = frm_principal.datainicial.text()
    data_final = frm_principal.datafinal.text()
    nome = frm_principal.comboBox_nome.currentText()
    advale = frm_principal.edt_advale.text() if frm_principal.edt_advale.text() else '00.00'
    dias = frm_principal.edt_dias.text()
    he = frm_principal.edt_he.text()
    subtotal = frm_principal.edt_subtotal.text()
    total = frm_principal.edt_total.text()
    vale = frm_principal.edt_vale.text() if frm_principal.edt_vale.text() else '00.00'
    vr = frm_principal.edt_vr.text()
    vt = frm_principal.edt_vt.text()

    # Define os títulos das colunas
    colunas = ['Data Inicial', 'Data Final', 'Nome', 'Dias TR',
               'HE', 'VT', 'VR', 'AD Vale', 'Vale', 'Subtotal', 'Total']

    # Adiciona os títulos das colunas na primeira linha se a planilha estiver vazia
    if ws.max_row == 1 and all([cell.value is None for cell in ws[1]]):
        for col_num, title in enumerate(colunas, 1):
            ws[get_column_letter(col_num) + '1'] = title

    # Adiciona os dados na próxima linha disponível
    ws.append([data_inicial, data_final, nome, dias, he,
              vt, vr, advale, vale, subtotal, total])

    # Salva o arquivo
    wb.save(arquivos_xlsx)

    # Limpando tela Registro
    frm_principal.edt_advale.setText('')
    frm_principal.edt_dias.setText('')
    frm_principal.edt_he.setText('')
    frm_principal.edt_subtotal.setText('')
    frm_principal.edt_total.setText('')
    frm_principal.edt_vale.setText('')
    frm_principal.edt_vr.setText('')
    frm_principal.edt_vt.setText('')


def limpardadosRegistro():
    # Caminho para o arquivo existente
    arquivos_xlsx = r'.\dados\RegistrosColaboradores.xlsx'
    # Carrega o arquivo Excel
    workbook = openpyxl.load_workbook(arquivos_xlsx)

    # Seleciona a aba especificada
    sheet = workbook['Registros']

    # Obtém o número de linhas com dados
    max_row = sheet.max_row

    # Exclui as linhas a partir da segunda linha
    sheet.delete_rows(2, max_row-1)

    # Salva o arquivo
    workbook.save(arquivos_xlsx)


def cadastroColaborador():
    nomecompleto = frm_principal.comboBox_nomeCad.currentText()
    funcao = frm_principal.comboBox_funcaoColab.currentText()
    cpf = frm_principal.edt_cpfCad.text()
    rg = frm_principal.edt_rgCad.text()
    cnh = frm_principal.edt_cnhCad.text()

    # Verifica se o campo nomecompleto não está vazio
    if nomecompleto.strip() != "":
        try:
            banco = sqlite3.connect(r'.\dados\banco_cadastro.db')
            cursor = banco.cursor()
            # Verifica se o colaborador já existe pelo nome completo
            cursor.execute(
                f"SELECT * FROM cadastro_colaborador WHERE nome_completo = '{nomecompleto}';")
            dados = cursor.fetchall()
            if dados != nomecompleto.strip():
                # Inserir novos dados na tabela
                cursor.execute("INSERT INTO cadastro_colaborador (nome_completo, funcao, cpf, rg, cnh) VALUES (?, ?, ?, ?, ?)", (
                    nomecompleto, funcao, cpf, rg, cnh))
                banco.commit()
                QMessageBox.information(
                    frm_principal, "Aviso", "Colaborador cadastrado com sucesso")
            else:
                cursor.execute("UPDATE cadastro_colaborador SET nome_completo = ?, funcao = ?, cpf = ?, rg = ? WHERE nome_completo = ?", (
                    nomecompleto, funcao, cpf, rg, nomecompleto))
                banco.commit()
                QMessageBox.information(
                    frm_principal, "Aviso", "Colaborador atualizado com sucesso")
        except sqlite3.Error as erro:
            QMessageBox.critical(frm_principal, "ERRO",
                                 f"{erro} ao inserir os dados")
        finally:
            # Fecha a conexão com o banco de dados
            banco.close()
    else:
        QMessageBox.warning(frm_principal, "Aviso",
                            "O campo 'Nome Completo' não pode estar vazio.")


def excluirColaborador():
    nomecompleto = frm_principal.comboBox_nomeCad.currentText()
    banco = sqlite3.connect(r'.\dados\banco_cadastro.db')
    cursor = banco.cursor()
    try:
        cursor.execute(
            "DELETE FROM cadastro_colaborador WHERE nome_completo = ?", (nomecompleto,))
        banco.commit()
        # Verifica se a função foi realmente excluída
        if cursor.rowcount > 0:
            QMessageBox.information(
                frm_principal, "Aviso", "Colaborador excluída com sucesso.")
    except sqlite3.Error as erro:
        print("Erro ao excluir a função: ", erro)
        QMessageBox.critical(frm_principal, "ERRO",
                             "Erro ao excluir a função.")
    finally:
        # Fecha a conexão com o banco de dados
        banco.close()


def cadastro_e_adicionar_taxas():
    # Captura os dados do formulário
    diaria = str(frm_principal.edt_diariaCad.text().replace(',', '.'))  # 85,00
    hextra = str(frm_principal.edt_hextraCad.text().replace(',', '.'))  # 16,00
    vtrans = str(
        frm_principal.edt_vtranspCad.text().replace(',', '.'))  # 12,30
    vresf = str(frm_principal.edt_vrefCad.text().replace(',', '.'))  # 17,00
    try:
        banco = sqlite3.connect(r'.\dados\banco_cadastro.db')
        cursor = banco.cursor()
        cursor.execute("select * from tabela")
        dados = cursor.fetchall()
        tables = len(dados)
        if tables == 0:
            # inserir dados na tabela
            cursor.execute("INSERT INTO tabela VALUES(NULL,'" +
                           diaria+"','"+hextra+"','"+vtrans+"','"+vresf+"')")
            banco.commit()
            banco.close()
            QMessageBox.information(
                frm_principal, "Aviso", "Tabela cadastrado com sucesso")
        else:
            cursor.execute("UPDATE tabela SET diaria = '"+diaria+"', hextra = '" +
                           hextra+"',vtransp = '"+vtrans+"', vref = '"+vresf+"'")
            banco.commit()
            banco.close()
            QMessageBox.information(
                frm_principal, "Aviso", "Tabela atualizado com sucesso")
    except sqlite3.Error as erro:
        print("Erro ao inserir os dados: ", erro)
        QMessageBox.about(frm_principal, "ERRO", "Erro ao inserir os dados")
    banco.close()


def preencherComboBoxFuncao():
    # Caminho para o arquivo do banco de dados
    database = r'.\dados\banco_cadastro.db'
    # Conecta ao banco de dados
    banco = sqlite3.connect(database)
    # Cria um cursor para executar operações no banco de dados
    cursor = banco.cursor()
    # Seleciona todas as descrições da tabela 'funcoes'
    cursor.execute("SELECT descricao FROM funcoes")
    # Recupera todos os resultados
    funcoes = cursor.fetchall()
    # Lista para armazenar as descrições
    lista_funcoes = [funcao[0] for funcao in funcoes]
    # Fecha a conexão com o banco de dados
    banco.close()
    return lista_funcoes


def cadastroFuncao():
    funcao = frm_principal.comboBox_funcaoCad.currentText()

    banco = sqlite3.connect(r'.\dados\banco_cadastro.db')
    cursor = banco.cursor()
    # Verifica se a descrição já existe na tabela 'funcoes'
    cursor.execute("SELECT id FROM funcoes WHERE descricao = ?", (funcao,))
    dados = cursor.fetchone()

    try:
        if dados is None:
            # Insere a nova função na tabela 'funcoes' se ela não existir
            cursor.execute(
                "INSERT INTO funcoes (descricao) VALUES (?)", (funcao,))
            banco.commit()
            QMessageBox.information(
                frm_principal, "Aviso", "Função cadastrada com sucesso.")
    except sqlite3.Error as erro:
        print("Erro ao inserir os dados: ", erro)
        QMessageBox.about(frm_principal, "ERRO", "Erro ao inserir os dados")
    banco.close()
    frm_principal.show()


def excluirFuncao():
    # Obtém o texto do item selecionado no comboBox
    funcao = frm_principal.comboBox_funcaoCad.currentText()
    # Conecta ao banco de dados
    banco = sqlite3.connect(r'.\dados\banco_cadastro.db')
    cursor = banco.cursor()
    try:
        # Exclui a função da tabela 'funcoes' com base na descrição
        cursor.execute("DELETE FROM funcoes WHERE descricao = ?", (funcao,))
        banco.commit()
        # Verifica se a função foi realmente excluída
        if cursor.rowcount > 0:
            QMessageBox.information(
                frm_principal, "Aviso", "Função excluída com sucesso.")
        else:
            QMessageBox.warning(frm_principal, "Aviso",
                                "Função não encontrada.")
    except sqlite3.Error as erro:
        QMessageBox.critical(frm_principal, "ERRO",
                             f"{erro} ao excluir a função.")
    finally:
        # Fecha a conexão com o banco de dados
        banco.close()


def preencherComboBoxRegitro():
    # Caminho para o arquivo do banco de dados
    database = r'.\dados\banco_cadastro.db'
    # Conecta ao banco de dados
    banco = sqlite3.connect(database)
    # Cria um cursor para executar operações no banco de dados
    cursor = banco.cursor()
    # Seleciona todas as descrições da tabela 'funcoes'
    cursor.execute("SELECT nome_completo FROM cadastro_colaborador")
    # Recupera todos os resultados
    colaboradores = cursor.fetchall()
    # Lista para armazenar as descrições
    lista_colaboradores = [colaborador[0] for colaborador in colaboradores]
    # Fecha a conexão com o banco de dados
    banco.close()
    return lista_colaboradores


def calcularRegistro():
    # +
    dias = str(frm_principal.edt_dias.text()).replace(',', '.')
    he = str(frm_principal.edt_he.text()).replace(',', '.')
    vr = str(frm_principal.edt_vr.text()).replace(',', '.')
    vt = str(frm_principal.edt_vt.text()).replace(',', '.')
    # -
    advale = str(frm_principal.edt_advale.text()).replace(',', '.')
    vale = str(frm_principal.edt_vale.text()).replace(',', '.')
    # tabela
    diaria = str(frm_principal.edt_diaria.text()).replace(',', '.')
    hextra = str(frm_principal.edt_hextra.text()).replace(',', '.')
    vtransp = str(frm_principal.edt_vtransp.text()).replace(',', '.')
    vref = str(frm_principal.edt_vref.text()).replace(',', '.')
    # calcular
    dias1 = float(dias) if dias else 0.0
    he1 = float(he) if he else 0.0
    vr1 = float(vr) if vr else 0.0
    vt1 = float(vt) if vt else 0.0
    advale1 = float(advale) if advale else 0.0
    vale1 = float(vale) if vale else 0.0
    diaria1 = float(diaria) if diaria else 0.0
    hextra1 = float(hextra) if hextra else 0.0
    vref1 = float(vref) if vref else 0.0
    vtransp1 = float(vtransp) if vtransp else 0.0
    tdias = (dias1 * diaria1)
    the = (he1 * hextra1)
    tvr = (vr1 * vref1)
    tvt = (vt1 * vtransp1)
    sobtotal = (tdias + the + tvr + tvt)
    subt = (advale1 + vale1)
    total = (sobtotal - subt)
    # Mostrar resultado
    frm_principal.edt_subtotal.setText("{:.2f}".format(sobtotal))
    frm_principal.edt_total.setText("{:.2f}".format(total))


def gerarWord():
    pass


def atualizarInterface():
    nome_arquivo = 'RegistrosColaboradores.xlsx'
    arquivo_xlsx = os.path.join(pasta_dados, nome_arquivo)
    if not os.path.isfile(arquivo_xlsx):
        gerarPlanilha()
    # Atualiza o comboBox com funções
    frm_principal.comboBox_funcaoCad.clear()
    frm_principal.comboBox_funcaoCad.addItems(preencherComboBoxFuncao())
    frm_principal.comboBox_funcaoColab.clear()
    frm_principal.comboBox_funcaoColab.addItems(preencherComboBoxFuncao())
    frm_principal.comboBox_nomeCad.clear()
    frm_principal.comboBox_nomeCad.addItems(preencherComboBoxRegitro())
    frm_principal.comboBox_nome.clear()
    frm_principal.comboBox_nome.addItems(preencherComboBoxRegitro())

    # Atualiza os campos de texto com as taxas mais recentes
    try:
        banco = sqlite3.connect(r'.\dados\banco_cadastro.db')
        cursor = banco.cursor()
        cursor.execute("SELECT * FROM tabela")
        dados_lidos = cursor.fetchall()
        frm_principal.edt_diariaCad.setText(
            str('%.2f' % dados_lidos[0][1]).replace('.', ','))
        frm_principal.edt_hextraCad.setText(
            str('%.2f' % dados_lidos[0][2]).replace('.', ','))
        frm_principal.edt_vtranspCad.setText(
            str('%.2f' % dados_lidos[0][3]).replace('.', ','))
        frm_principal.edt_vrefCad.setText(
            str('%.2f' % dados_lidos[0][4]).replace('.', ','))
        frm_principal.edt_diaria.setText(
            str('%.2f' % dados_lidos[0][1]).replace('.', ','))
        frm_principal.edt_hextra.setText(
            str('%.2f' % dados_lidos[0][2]).replace('.', ','))
        frm_principal.edt_vtransp.setText(
            str('%.2f' % dados_lidos[0][3]).replace('.', ','))
        frm_principal.edt_vref.setText(
            str('%.2f' % dados_lidos[0][4]).replace('.', ','))
    except sqlite3.Error as erro:
        print("Erro ao atualizar os dados: ", erro)
    atualizarListaRegistro()
    # limpar tela cadastro Colaborador
    frm_principal.edt_cpfCad.setText('')
    frm_principal.edt_rgCad.setText('')
    frm_principal.edt_cnhCad.setText('')


def atualizarListaRegistro():
    # Carrega o arquivo Excel
    arquivos_xlsx = r'.\dados\RegistrosColaboradores.xlsx'
    workbook = openpyxl.load_workbook(arquivos_xlsx)

    # Seleciona a primeira aba ativa do workbook
    sheet = workbook.active

    # Configura o QTableWidget com o número de linhas e colunas
    # Subtrai 1 para não contar os cabeçalhos
    frm_principal.tableWidget.setRowCount(sheet.max_row - 1)
    frm_principal.tableWidget.setColumnCount(sheet.max_column)

    # Define os cabeçalhos das colunas com os valores da primeira linha
    headers = [cell.value for cell in next(sheet.iter_rows(max_row=1))]
    frm_principal.tableWidget.setHorizontalHeaderLabels(headers)

    # Preenche o QTableWidget com os dados, começando da segunda linha
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=0):
        for column_index, value in enumerate(row, start=0):
            item = QTableWidgetItem(str(value))
            frm_principal.tableWidget.setItem(row_index, column_index, item)

    # Ajusta o tamanho das colunas para se ajustar ao conteúdo
    frm_principal.tableWidget.resizeColumnsToContents()


if __name__ == '__main__':
    App = QtWidgets.QApplication([])
    frm_principal = uic.loadUi(r'.\frms\frm_principal.ui')
    # Conecta os eventos aos métodos correspondentes
    frm_principal.btn_GerarTabela.clicked.connect(gerarPlanilha)
    frm_principal.btn_salvarTabela.clicked.connect(
        lambda: [cadastro_e_adicionar_taxas(), atualizarInterface()])
    frm_principal.btn_salvarFucao.clicked.connect(
        lambda: [cadastroFuncao(), atualizarInterface()])
    frm_principal.btn_excluirFucao.clicked.connect(
        lambda: [excluirFuncao(), atualizarInterface()])
    frm_principal.btn_salvarRegistro.clicked.connect(
        lambda: [salvarRegistro(), atualizarInterface()])
    frm_principal.btn_salvarCadastro.clicked.connect(
        lambda: [cadastroColaborador(), atualizarInterface()])
    frm_principal.btn_calcular.clicked.connect(
        lambda: [calcularRegistro(), atualizarInterface()])
    frm_principal.btn_excluirCad.clicked.connect(
        lambda: [excluirColaborador(), atualizarInterface()])
    frm_principal.btn_excluirTabela.clicked.connect(
        lambda: [limpardadosRegistro(), atualizarInterface()])
    # Atualiza a interface ao iniciar o aplicativo
    atualizarInterface()
    frm_principal.show()
    App.exec()
